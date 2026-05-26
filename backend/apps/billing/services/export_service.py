"""
Export service for PEDIACORE billing.

Provides XLSX export of payment records using openpyxl.
"""

from __future__ import annotations

import io
import logging
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from django.db.models import QuerySet

    from apps.billing.models import Payment

logger = logging.getLogger(__name__)

EXPORT_COLUMNS = [
    "Date",
    "Patient",
    "Service",
    "Amount",
    "Currency",
    "Status",
    "Method",
    "External ID",
    "Invoice Number",
    "Notes",
]


def export_payments_xlsx(queryset: QuerySet) -> bytes:
    """
    Export a Payment queryset to XLSX format.

    Columns: date, patient, service, amount, currency, status, method,
             external_id, invoice_number, notes.

    Args:
        queryset: A Payment QuerySet, pre-filtered by the caller.

    Returns:
        bytes: The XLSX file content as bytes.

    Raises:
        ImportError: If openpyxl is not installed. Callers should handle this.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        logger.error("openpyxl is not installed — cannot export XLSX: %s", exc)
        raise

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payments"

    # Header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")

    for col_idx, header in enumerate(EXPORT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill

    # Data rows
    payments = queryset.select_related(
        "patient",
        "appointment__service",
        "invoice",
    )

    for row_idx, payment in enumerate(payments, start=2):
        patient_name = f"{payment.patient.first_name} {payment.patient.last_name}"

        service_name = ""
        if payment.appointment and payment.appointment.service:
            service_name = payment.appointment.service.name

        invoice_number = ""
        try:
            invoice_number = payment.invoice.invoice_number
        except Exception:
            pass

        paid_at = payment.paid_at
        date_str = paid_at.strftime("%Y-%m-%d %H:%M") if paid_at else str(payment.created_at.date())

        row_data = [
            date_str,
            patient_name,
            service_name,
            float(payment.amount),
            payment.currency,
            payment.get_status_display(),
            payment.get_payment_method_display(),
            payment.external_id,
            invoice_number,
            payment.notes,
        ]

        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-fit column widths (approximate)
    for col_idx in range(1, len(EXPORT_COLUMNS) + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 18

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    logger.info("Exported %d payments to XLSX", payments.count())
    return buffer.getvalue()
