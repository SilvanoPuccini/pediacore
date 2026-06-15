"""
Views for the PEDIACORE billing app.

Endpoints:
  - PaymentViewSet: CRUD for payments + MercadoPago preference + XLSX export
  - InvoiceViewSet: list/detail for invoices + PDF download
  - PaymentProviderViewSet: admin CRUD for payment provider configs
  - MercadoPagoWebhookView: standalone webhook endpoint for MP notifications
"""

from __future__ import annotations

import logging
from datetime import datetime, date
from decimal import Decimal

import mercadopago
from django.conf import settings
from django.contrib.auth import get_user_model
from django.db import transaction
from django.db.models import Sum
from django.db.models.functions import TruncMonth
from django.http import HttpResponse
from django.utils import timezone
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.request import Request
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.billing.models import Invoice, MonthlyExpense, Payment, PaymentProvider
from apps.billing.serializers import (
    InvoiceDetailSerializer,
    InvoiceListSerializer,
    MonthlyExpenseSerializer,
    PaymentCreateSerializer,
    PaymentDetailSerializer,
    PaymentListSerializer,
    PaymentProviderSerializer,
    PaymentUpdateSerializer,
    ReceiptUploadSerializer,
    TransferConfirmSerializer,
    TransferRejectSerializer,
)
from apps.billing.services.export_service import export_payments_xlsx
from apps.billing.services.invoice_service import create_invoice_for_payment, generate_invoice_pdf
from apps.billing.services.payment_strategy import MercadoPagoStrategy, get_payment_strategy
from apps.core.pagination import StandardPagination
from apps.core.permissions import IsDoctor, IsTutor
from apps.notifications.services.email_service import (
    send_appointment_confirmation,
    send_payment_receipt,
    send_transfer_confirmed,
    send_transfer_receipt_uploaded,
    send_transfer_rejected,
)
from apps.scheduling.models import Appointment
from apps.scheduling.services.token_service import create_tokens_for_appointment

User = get_user_model()

logger = logging.getLogger(__name__)


def _build_token_urls(tokens: list) -> dict:
    """
    Build a token_urls dict suitable for send_appointment_confirmation.

    Maps each AppointmentToken's action to its frontend URL.
    CONFIRM and CANCEL point to the same resolve page; RESCHEDULE goes to
    the reschedule form.
    """
    frontend_url = getattr(settings, "FRONTEND_URL", "https://estefipediatra.com")
    url_map: dict[str, str | None] = {"confirm": None, "cancel": None, "reschedule": None}

    for token in tokens:
        if token.action == "CONFIRM":
            url_map["confirm"] = f"{frontend_url}/a/{token.token}/"
        elif token.action == "CANCEL":
            url_map["cancel"] = f"{frontend_url}/a/{token.token}/"
        elif token.action == "RESCHEDULE":
            url_map["reschedule"] = f"{frontend_url}/a/{token.token}/reschedule"

    return url_map


class PaymentViewSet(viewsets.ModelViewSet):
    """
    Payments CRUD + MercadoPago preference creation + webhook + XLSX export.

    Access control:
      - DOCTOR: sees all payments, can create/update
      - TUTOR: sees only payments for their linked patients
      - Anonymous: no access
    """

    pagination_class = StandardPagination
    http_method_names = ["get", "post", "patch", "head", "options"]

    def get_permissions(self):
        if self.action in ("update", "partial_update", "export"):
            return [IsDoctor()]
        return [IsAuthenticated()]

    def get_serializer_class(self):
        if self.action == "create":
            return PaymentCreateSerializer
        if self.action in ("update", "partial_update"):
            return PaymentUpdateSerializer
        if self.action == "retrieve":
            return PaymentDetailSerializer
        return PaymentListSerializer

    def get_queryset(self):
        user = self.request.user

        qs = Payment.objects.select_related(
            "practice", "patient", "appointment",
            "appointment__service", "appointment__location", "paid_by",
            "invoice",
        )

        if user.role == User.TUTOR:
            from apps.patients.models import TutorPatient

            patient_ids = TutorPatient.objects.filter(tutor=user).values_list("patient_id", flat=True)
            qs = qs.filter(patient_id__in=patient_ids)
        elif user.role != User.DOCTOR:
            qs = qs.none()

        # Optional filters
        status_param = self.request.query_params.get("status")
        if status_param:
            qs = qs.filter(status=status_param)

        patient_id_param = self.request.query_params.get("patient_id")
        if patient_id_param:
            qs = qs.filter(patient_id=patient_id_param)

        payment_method_param = self.request.query_params.get("payment_method")
        if payment_method_param:
            qs = qs.filter(payment_method=payment_method_param)

        receipt_uploaded = self.request.query_params.get("receipt_uploaded")
        if receipt_uploaded == "true":
            qs = qs.exclude(receipt_uploaded_at__isnull=True)
        elif receipt_uploaded == "false":
            qs = qs.filter(receipt_uploaded_at__isnull=True)

        return qs

    def perform_create(self, serializer):
        """
        Inject the payment amount from the appointment's service price for
        TUTOR requests, so tutors cannot specify an arbitrary amount.
        DOCTOR requests may include amount directly (e.g. manual cash payments
        without an appointment).
        """
        from rest_framework.exceptions import ValidationError

        request = self.request
        appointment = serializer.validated_data.get("appointment")

        if request.user.role == User.TUTOR:
            if not appointment or not appointment.service:
                raise ValidationError(
                    {"appointment": "Tutors must provide a valid appointment to create a payment."}
                )
            amount = appointment.service.price_clp
            serializer.save(amount=amount)
        else:
            # DOCTOR: amount must be present (either from appointment or explicit input)
            if not serializer.validated_data.get("amount") and appointment and appointment.service:
                serializer.save(amount=appointment.service.price_clp)
            else:
                serializer.save()

    @action(detail=True, methods=["post"], url_path="create-preference")
    def create_preference(self, request: Request, pk=None) -> Response:
        """
        Create a payment preference with the configured provider.

        For MercadoPago, returns an init_point URL to redirect the user
        to complete payment on MP's hosted checkout page.
        """
        payment = self.get_object()

        if payment.status == Payment.COMPLETED:
            return Response(
                {"detail": "Payment is already completed."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        strategy = get_payment_strategy(payment.payment_method)
        result = strategy.create_preference(payment)

        return Response(result, status=status.HTTP_200_OK)

    @action(detail=False, methods=["get"], permission_classes=[IsDoctor])
    def export(self, request: Request) -> HttpResponse:
        """
        Export all payments as an XLSX file (doctor only).

        Applies the same queryset filters as list (status, patient_id).
        """
        queryset = self.get_queryset()

        try:
            xlsx_bytes = export_payments_xlsx(queryset)
        except ImportError:
            return Response(
                {"detail": "openpyxl is not installed. Cannot export XLSX."},
                status=status.HTTP_501_NOT_IMPLEMENTED,
            )

        response = HttpResponse(
            content=xlsx_bytes,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="payments.xlsx"'
        return response

    @action(
        detail=True,
        methods=["post"],
        url_path="upload-receipt",
        permission_classes=[IsTutor],
        parser_classes=[MultiPartParser],
    )
    def upload_receipt(self, request: Request, pk=None) -> Response:
        """
        Upload a bank transfer receipt file for a pending transfer payment.

        Only the tutor who owns the payment (paid_by) may upload.
        Payment must be PENDING with payment_method=TRANSFER.
        Accepts: PDF, JPG, JPEG, PNG. Max size: 10 MB.
        """
        payment = self.get_object()

        # Ownership check: tutor must be the payer
        if payment.paid_by != request.user:
            return Response(
                {"detail": "You do not have permission to upload a receipt for this payment."},
                status=status.HTTP_403_FORBIDDEN,
            )

        if payment.payment_method != Payment.TRANSFER:
            return Response(
                {"detail": "Receipt upload is only valid for bank transfer payments."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if payment.status != Payment.PENDING:
            return Response(
                {"detail": "Receipt can only be uploaded for pending payments."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        serializer = ReceiptUploadSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        receipt_file = serializer.validated_data["receipt"]

        payment.receipt_file = receipt_file
        payment.receipt_uploaded_at = timezone.now()
        payment.save(update_fields=["receipt_file", "receipt_uploaded_at", "updated_at"])

        try:
            send_transfer_receipt_uploaded(payment)
        except Exception as exc:
            logger.error(
                "upload_receipt: failed to send notification for Payment #%s: %s",
                payment.pk,
                exc,
            )

        # Fire-and-forget: analyse the receipt with Gemini OCR in the background.
        # The upload succeeds regardless of whether the task is enqueued.
        try:
            from django_q.tasks import async_task

            async_task(
                "apps.billing.services.ocr_service.analyze_receipt_with_gemini",
                payment.id,
            )
            logger.info("upload_receipt: OCR task enqueued for Payment #%s", payment.pk)
        except Exception as exc:
            logger.error(
                "upload_receipt: could not enqueue OCR task for Payment #%s: %s",
                payment.pk,
                exc,
            )

        return Response(
            {
                "status": "pending",
                "receipt_uploaded_at": payment.receipt_uploaded_at,
            },
            status=status.HTTP_200_OK,
        )

    @action(
        detail=True,
        methods=["post"],
        url_path="confirm-transfer",
        permission_classes=[IsDoctor],
    )
    def confirm_transfer(self, request: Request, pk=None) -> Response:
        """
        Confirm a bank transfer payment (doctor only).

        Transitions Payment → COMPLETED, Appointment → CONFIRMED.
        Creates an Invoice and sends confirmation email to the tutor.
        """
        payment = self.get_object()

        if payment.payment_method != Payment.TRANSFER:
            return Response(
                {"detail": "confirm-transfer is only valid for bank transfer payments."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if payment.status != Payment.PENDING:
            return Response(
                {"detail": "Only pending transfer payments can be confirmed."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        serializer = TransferConfirmSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        notes = serializer.validated_data.get("notes", "")

        with transaction.atomic():
            payment.status = Payment.COMPLETED
            payment.paid_at = timezone.now()
            if notes:
                payment.notes = notes
            payment.save(update_fields=["status", "paid_at", "notes", "updated_at"])

            appointment = payment.appointment
            if appointment and appointment.status == Appointment.PENDING:
                appointment.status = Appointment.CONFIRMED
                appointment.confirmed_at = timezone.now()
                appointment.save(update_fields=["status", "confirmed_at", "updated_at"])

        # Post-confirmation pipeline (non-blocking)
        try:
            create_invoice_for_payment(payment)
            logger.info("confirm_transfer: Invoice created for Payment #%s", payment.pk)
        except Exception as exc:
            logger.error(
                "confirm_transfer: invoice creation failed for Payment #%s: %s",
                payment.pk,
                exc,
            )

        try:
            if hasattr(payment, "invoice"):
                generate_invoice_pdf(payment.invoice)
            else:
                logger.warning("confirm_transfer: no invoice to generate PDF for Payment #%s", payment.pk)
        except Exception as exc:
            logger.error(
                "confirm_transfer: PDF generation failed for Payment #%s: %s",
                payment.pk,
                exc,
            )

        # ── Generate Zoom meeting if applicable ────────────────────────────
        try:
            if appointment and appointment.is_online and appointment.call_platform == "ZOOM":
                from apps.scheduling.services.zoom_service import create_zoom_meeting

                meeting_start = datetime.combine(
                    appointment.scheduled_date, appointment.start_time
                )
                join_url = create_zoom_meeting(
                    topic=f"Consulta — {appointment.patient.full_name}",
                    start_time=meeting_start,
                    duration_minutes=appointment.service.duration_minutes,
                )
                appointment.meeting_link = join_url
                appointment.save(update_fields=["meeting_link", "updated_at"])
        except Exception as exc:
            logger.error(
                "confirm_transfer: Zoom meeting creation failed for Appointment #%s: %s",
                appointment.pk if appointment else "—",
                exc,
            )

        # ── Create tokens + send confirmation emails with action buttons ────
        token_urls = None
        try:
            if appointment:
                tokens = create_tokens_for_appointment(appointment)
                token_urls = _build_token_urls(tokens)
        except Exception as exc:
            logger.error(
                "confirm_transfer: token creation failed for Appointment #%s: %s",
                appointment.pk if appointment else "—",
                exc,
            )

        try:
            send_transfer_confirmed(payment)
        except Exception as exc:
            logger.error(
                "confirm_transfer: send_transfer_confirmed failed for Payment #%s: %s",
                payment.pk,
                exc,
            )

        try:
            if appointment:
                send_appointment_confirmation(appointment, token_urls=token_urls)
        except Exception as exc:
            logger.error(
                "confirm_transfer: send_appointment_confirmation failed for Appointment #%s: %s",
                appointment.pk if appointment else "—",
                exc,
            )

        return Response({"status": "completed"}, status=status.HTTP_200_OK)

    @action(
        detail=True,
        methods=["post"],
        url_path="process-card",
        permission_classes=[IsTutor],
    )
    def process_card(self, request: Request, pk=None) -> Response:
        """
        Process an inline card payment using a token from the Payment Brick.

        The frontend's CardPayment/Payment Brick tokenizes the card and sends
        the token + payment details here. We call the MercadoPago API to create
        the actual payment, then confirm the appointment.

        Expected payload:
            token, payment_method_id, issuer_id, installments,
            payer.email, payer.identification.type, payer.identification.number
        """
        payment = self.get_object()

        # Ownership check
        if payment.paid_by != request.user:
            return Response(
                {"detail": "You do not have permission to process this payment."},
                status=status.HTTP_403_FORBIDDEN,
            )

        if payment.status != Payment.PENDING:
            return Response(
                {"detail": "Only pending payments can be processed."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Extract token data from request
        data = request.data
        token = data.get("token")
        if not token:
            return Response(
                {"detail": "Card token is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        from decimal import Decimal

        amount = int(Decimal(str(payment.amount)))

        # Resolve service description
        service_name = "Consulta pediátrica"
        if payment.appointment and payment.appointment.service:
            service_name = payment.appointment.service.name

        # Build payer info
        payer_data = data.get("payer", {})
        payer_email = payer_data.get("email", request.user.email)
        identification = payer_data.get("identification", {})

        payment_data = {
            "transaction_amount": amount,
            "token": token,
            "description": service_name,
            "installments": int(data.get("installments", 1)),
            "payment_method_id": data.get("payment_method_id", ""),
            "payer": {
                "email": payer_email,
            },
        }

        # issuer_id must be an integer for MP API — omit if empty/missing
        raw_issuer = data.get("issuer_id")
        if raw_issuer:
            try:
                payment_data["issuer_id"] = int(raw_issuer)
            except (ValueError, TypeError):
                pass

        # Add identification if provided
        if identification.get("type") and identification.get("number"):
            payment_data["payer"]["identification"] = {
                "type": identification["type"],
                "number": identification["number"],
            }

        # Call MercadoPago API to create the payment
        access_token = getattr(settings, "MERCADOPAGO_ACCESS_TOKEN", "")
        sdk = mercadopago.SDK(access_token)

        logger.info(
            "process_card: Payment #%s — token=%s...%s access_token=%s...%s payload=%s",
            payment.pk,
            token[:8] if token else "EMPTY",
            token[-4:] if token and len(token) > 8 else "",
            access_token[:20] if access_token else "EMPTY",
            access_token[-10:] if access_token else "",
            {k: v for k, v in payment_data.items() if k != "token"},
        )

        try:
            mp_response = sdk.payment().create(payment_data)
        except Exception as exc:
            logger.error("process_card: MP API call failed for Payment #%s: %s", payment.pk, exc)
            return Response(
                {"detail": "Error processing payment with MercadoPago."},
                status=status.HTTP_502_BAD_GATEWAY,
            )

        mp_status_code = mp_response.get("status")
        mp_result = mp_response.get("response", {})
        mp_payment_status = mp_result.get("status", "")
        mp_payment_id = str(mp_result.get("id", ""))

        logger.info(
            "process_card: MP response status=%s mp_payment_status=%s mp_id=%s for Payment #%s",
            mp_status_code,
            mp_payment_status,
            mp_payment_id,
            payment.pk,
        )

        if mp_status_code not in (200, 201):
            error_message = mp_result.get("message", "Payment was rejected.")
            logger.error(
                "process_card: MP REJECTED Payment #%s — http=%s cause=%s full_response=%s",
                payment.pk, mp_status_code, mp_result.get("cause", []), mp_response,
            )
            # Upstream MP failure (5xx) → 502; client-side rejection → 400
            http_status = (
                status.HTTP_502_BAD_GATEWAY
                if mp_status_code >= 500
                else status.HTTP_400_BAD_REQUEST
            )
            return Response(
                {"detail": error_message, "mp_status": mp_payment_status},
                status=http_status,
            )

        # Payment approved — update records
        if mp_payment_status == "approved":
            with transaction.atomic():
                payment.status = Payment.COMPLETED
                payment.paid_at = timezone.now()
                payment.external_id = mp_payment_id
                payment.external_status = mp_payment_status
                payment.metadata.update({
                    "provider": "mercadopago",
                    "mp_payment_id": mp_payment_id,
                    "processing_mode": "card_inline",
                })
                payment.save(update_fields=[
                    "status", "paid_at", "external_id", "external_status",
                    "metadata", "updated_at",
                ])

                appointment = payment.appointment
                if appointment and appointment.status in (Appointment.PENDING, Appointment.HOLD):
                    appointment.status = Appointment.CONFIRMED
                    appointment.confirmed_at = timezone.now()
                    appointment.save(update_fields=["status", "confirmed_at", "updated_at"])

            # Post-confirmation pipeline (non-blocking)
            try:
                create_invoice_for_payment(payment)
            except Exception as exc:
                logger.error("process_card: invoice creation failed for Payment #%s: %s", payment.pk, exc)

            try:
                if hasattr(payment, "invoice"):
                    generate_invoice_pdf(payment.invoice)
            except Exception as exc:
                logger.error("process_card: PDF generation failed for Payment #%s: %s", payment.pk, exc)

            try:
                send_payment_receipt(payment)
            except Exception as exc:
                logger.error("process_card: send_payment_receipt failed for Payment #%s: %s", payment.pk, exc)

            # ── Generate Zoom meeting if applicable ────────────────────────────────
            try:
                if appointment and appointment.is_online and appointment.call_platform == "ZOOM":
                    from apps.scheduling.services.zoom_service import create_zoom_meeting

                    meeting_start = datetime.combine(
                        appointment.scheduled_date, appointment.start_time
                    )
                    join_url = create_zoom_meeting(
                        topic=f"Consulta — {appointment.patient.full_name}",
                        start_time=meeting_start,
                        duration_minutes=appointment.service.duration_minutes,
                    )
                    appointment.meeting_link = join_url
                    appointment.save(update_fields=["meeting_link", "updated_at"])
                    logger.info(
                        "process_card: Zoom meeting created for Appointment #%s: %s",
                        appointment.pk,
                        join_url,
                    )
            except Exception as exc:
                logger.error(
                    "process_card: Zoom meeting creation failed for Appointment #%s: %s",
                    appointment.pk if appointment else "—",
                    exc,
                )

            # ── Create tokens + send confirmation email with action buttons ────────
            try:
                if appointment:
                    tokens = create_tokens_for_appointment(appointment)
                    token_urls = _build_token_urls(tokens)
                    send_appointment_confirmation(appointment, token_urls=token_urls)
            except Exception as exc:
                logger.error("process_card: send_appointment_confirmation failed: %s", exc)

            return Response(
                {"status": "approved", "appointment_id": appointment.pk if appointment else None},
                status=status.HTTP_200_OK,
            )

        # Pending/in_process — store external_id but don't confirm yet
        payment.external_id = mp_payment_id
        payment.external_status = mp_payment_status
        payment.metadata.update({"provider": "mercadopago", "mp_payment_id": mp_payment_id})
        payment.save(update_fields=["external_id", "external_status", "metadata", "updated_at"])

        return Response(
            {"status": mp_payment_status, "detail": "Payment is being processed."},
            status=status.HTTP_202_ACCEPTED,
        )

    @action(
        detail=True,
        methods=["post"],
        url_path="reject-transfer",
        permission_classes=[IsDoctor],
    )
    def reject_transfer(self, request: Request, pk=None) -> Response:
        """
        Reject a bank transfer payment (doctor only).

        Requires a 'reason' in the request body.
        Transitions Payment → FAILED, Appointment → CANCELLED.
        Sends a rejection email to the tutor.
        """
        payment = self.get_object()

        if payment.payment_method != Payment.TRANSFER:
            return Response(
                {"detail": "reject-transfer is only valid for bank transfer payments."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if payment.status != Payment.PENDING:
            return Response(
                {"detail": "Only pending transfer payments can be rejected."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        serializer = TransferRejectSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        reason = serializer.validated_data["reason"]

        with transaction.atomic():
            payment.status = Payment.FAILED
            payment.save(update_fields=["status", "updated_at"])

            appointment = payment.appointment
            if appointment:
                appointment.status = Appointment.CANCELLED
                appointment.save(update_fields=["status", "updated_at"])

        try:
            send_transfer_rejected(payment, reason)
        except Exception as exc:
            logger.error(
                "reject_transfer: send_transfer_rejected failed for Payment #%s: %s",
                payment.pk,
                exc,
            )

        return Response({"status": "rejected"}, status=status.HTTP_200_OK)


class InvoiceViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Read-only viewset for invoices.

    Invoices are created automatically when a payment completes — they are
    never created directly via the API. Provides list, detail, and PDF download.
    """

    pagination_class = StandardPagination

    def get_permissions(self):
        return [IsAuthenticated()]

    def get_serializer_class(self):
        if self.action == "retrieve":
            return InvoiceDetailSerializer
        return InvoiceListSerializer

    def get_queryset(self):
        user = self.request.user

        qs = Invoice.objects.select_related("practice", "payment", "payment__patient")

        if user.role == User.TUTOR:
            from apps.patients.models import TutorPatient

            patient_ids = TutorPatient.objects.filter(tutor=user).values_list("patient_id", flat=True)
            qs = qs.filter(payment__patient_id__in=patient_ids)
        elif user.role != User.DOCTOR:
            qs = qs.none()

        return qs

    @action(detail=True, methods=["get"])
    def download(self, request: Request, pk=None) -> HttpResponse:
        """
        Download or generate the invoice PDF.

        If the PDF file does not exist yet, generates it on-the-fly.
        """
        invoice = self.get_object()

        try:
            if not invoice.pdf_file:
                pdf_bytes = generate_invoice_pdf(invoice)
            else:
                pdf_bytes = invoice.pdf_file.read()
        except Exception as exc:
            logger.error("Invoice PDF generation failed for #%s: %s", invoice.pk, exc)
            return Response(
                {"detail": "Failed to generate invoice PDF."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        response = HttpResponse(content=pdf_bytes, content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{invoice.invoice_number}.pdf"'
        return response


class MercadoPagoWebhookView(APIView):
    """
    Standalone webhook endpoint for MercadoPago payment notifications.

    URL: POST /api/v1/webhooks/mercadopago/
    Permission: AllowAny — MP sends requests without authentication.

    Pipeline on payment.approved:
      1. Validate X-Signature HMAC → 403 if invalid
      2. Ignore non-payment events → 200
      3. Fetch full payment from MP API by data.id
      4. Look up our Payment by external_reference (= Payment.pk)
      5. Idempotency: if already COMPLETED → 200 (skip)
      6. atomic: Payment→COMPLETED, Appointment→CONFIRMED
      7. create_invoice_for_payment() + send_payment_receipt() + send_appointment_confirmation()

    On payment.rejected:
      - Payment→FAILED; Appointment stays HOLD (will expire via background task)
    """

    permission_classes = [AllowAny]

    def post(self, request: Request) -> Response:
        """Handle a MercadoPago webhook notification."""
        # ── 1. Validate HMAC signature ────────────────────────────────────────
        webhook_secret = getattr(settings, "MERCADOPAGO_WEBHOOK_SECRET", "")
        data = request.data
        data_id = str(data.get("data", {}).get("id", ""))
        request_id = request.META.get("HTTP_X_REQUEST_ID", "")

        if webhook_secret:
            sig_valid = MercadoPagoStrategy.validate_webhook_signature(
                headers=request.META,
                data_id=data_id,
                request_id=request_id,
                secret=webhook_secret,
            )
            if not sig_valid:
                logger.warning(
                    "MercadoPago webhook: HMAC signature failed for data_id=%s — rejecting request",
                    data_id,
                )
                return Response({"detail": "Invalid signature."}, status=status.HTTP_403_FORBIDDEN)
        else:
            logger.info("MercadoPago webhook: MERCADOPAGO_WEBHOOK_SECRET not set, skipping signature validation")

        # ── 2. Ignore non-payment events ──────────────────────────────────────
        notification_type = data.get("type", "")
        if notification_type != "payment":
            logger.info("MercadoPago webhook: ignoring event type=%s", notification_type)
            return Response({"detail": "Event type ignored."}, status=status.HTTP_200_OK)

        if not data_id:
            logger.warning("MercadoPago webhook: missing data.id in payload")
            return Response({"detail": "Missing data.id."}, status=status.HTTP_400_BAD_REQUEST)

        # ── 3. Fetch full payment from MP API ─────────────────────────────────
        try:
            access_token = getattr(settings, "MERCADOPAGO_ACCESS_TOKEN", "")
            sdk = mercadopago.SDK(access_token)
            mp_response = sdk.payment().get(data_id)

            if mp_response.get("status") not in (200, 201):
                logger.error(
                    "MercadoPago webhook: failed to fetch payment data_id=%s status=%s",
                    data_id,
                    mp_response.get("status"),
                )
                return Response(
                    {"detail": "Failed to retrieve payment from MP."},
                    status=status.HTTP_502_BAD_GATEWAY,
                )

            mp_payment = mp_response["response"]
        except Exception as exc:
            logger.error("MercadoPago webhook: SDK error fetching payment: %s", exc, exc_info=True)
            return Response(
                {"detail": "Error fetching payment from MercadoPago."},
                status=status.HTTP_502_BAD_GATEWAY,
            )

        # ── 4. Look up our Payment by external_reference ──────────────────────
        external_reference = str(mp_payment.get("external_reference", ""))
        mp_payment_id = str(mp_payment.get("id", ""))
        mp_status = mp_payment.get("status", "")

        try:
            payment = Payment.objects.select_related("appointment", "appointment__service").get(pk=external_reference)
        except (Payment.DoesNotExist, ValueError):
            logger.warning(
                "MercadoPago webhook: no payment found for external_reference=%s",
                external_reference,
            )
            return Response({"detail": "Payment not found."}, status=status.HTTP_404_NOT_FOUND)

        # ── 5. Idempotency: already processed → skip ──────────────────────────
        if payment.status == Payment.COMPLETED:
            logger.info(
                "MercadoPago webhook: Payment #%s already COMPLETED, skipping (idempotent)",
                payment.pk,
            )
            return Response({"detail": "Already processed."}, status=status.HTTP_200_OK)

        # ── 6. Process based on MP status ─────────────────────────────────────
        if mp_status == "approved":
            try:
                with transaction.atomic():
                    # Update Payment
                    payment.status = Payment.COMPLETED
                    payment.external_id = mp_payment_id
                    payment.external_status = "approved"
                    payment.paid_at = timezone.now()
                    payment.save(update_fields=["status", "external_id", "external_status", "paid_at", "updated_at"])

                    # Update Appointment
                    appointment = payment.appointment
                    if appointment and appointment.status == Appointment.HOLD:
                        appointment.status = Appointment.CONFIRMED
                        appointment.confirmed_at = timezone.now()
                        appointment.save(update_fields=["status", "confirmed_at", "updated_at"])

                logger.info(
                    "MercadoPago webhook: Payment #%s COMPLETED, Appointment #%s CONFIRMED",
                    payment.pk,
                    appointment.pk if appointment else "—",
                )
            except Exception as exc:
                logger.error(
                    "MercadoPago webhook: DB error processing approval for Payment #%s: %s",
                    payment.pk,
                    exc,
                    exc_info=True,
                )
                return Response(
                    {"detail": "Error processing payment."},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                )

            # ── 7. Post-payment pipeline (non-blocking — failures logged, not raised) ──
            try:
                create_invoice_for_payment(payment)
                logger.info("MercadoPago webhook: Invoice created for Payment #%s", payment.pk)
            except Exception as exc:
                logger.error(
                    "MercadoPago webhook: invoice creation failed for Payment #%s: %s",
                    payment.pk,
                    exc,
                )

            try:
                send_payment_receipt(payment)
            except Exception as exc:
                logger.error(
                    "MercadoPago webhook: send_payment_receipt failed for Payment #%s: %s",
                    payment.pk,
                    exc,
                )

            # ── 7b. Generate Zoom meeting link if applicable ────────────────────
            try:
                if appointment and appointment.is_online and appointment.call_platform == "ZOOM":
                    from apps.scheduling.services.zoom_service import create_zoom_meeting

                    meeting_start = datetime.combine(
                        appointment.scheduled_date, appointment.start_time
                    )
                    join_url = create_zoom_meeting(
                        topic=f"Consulta — {appointment.patient.full_name}",
                        start_time=meeting_start,
                        duration_minutes=appointment.service.duration_minutes,
                    )
                    appointment.meeting_link = join_url
                    appointment.save(update_fields=["meeting_link", "updated_at"])
                    logger.info(
                        "Zoom meeting created for Appointment #%s: %s",
                        appointment.pk,
                        join_url,
                    )
            except Exception as exc:
                logger.error(
                    "Zoom meeting creation failed for Appointment #%s: %s",
                    appointment.pk if appointment else "—",
                    exc,
                )

            # Create tokens before sending confirmation so the email can include action links.
            token_urls = None
            try:
                if appointment:
                    tokens = create_tokens_for_appointment(appointment)
                    token_urls = _build_token_urls(tokens)
                    logger.info(
                        "MercadoPago webhook: tokens created for Appointment #%s",
                        appointment.pk,
                    )
            except Exception as exc:
                logger.error(
                    "MercadoPago webhook: token creation failed for Appointment #%s: %s",
                    appointment.pk if appointment else "—",
                    exc,
                )

            try:
                if appointment:
                    send_appointment_confirmation(appointment, token_urls=token_urls)
            except Exception as exc:
                logger.error(
                    "MercadoPago webhook: send_appointment_confirmation failed for Appointment #%s: %s",
                    appointment.pk if appointment else "—",
                    exc,
                )

        elif mp_status == "rejected":
            payment.status = Payment.FAILED
            payment.external_id = mp_payment_id
            payment.external_status = "rejected"
            payment.save(update_fields=["status", "external_id", "external_status", "updated_at"])
            logger.info(
                "MercadoPago webhook: Payment #%s FAILED (rejected by MP), Appointment stays HOLD",
                payment.pk,
            )

        else:
            logger.info(
                "MercadoPago webhook: unhandled MP status=%s for Payment #%s",
                mp_status,
                payment.pk,
            )

        return Response({"detail": "OK"}, status=status.HTTP_200_OK)


class TaxCalculatorView(APIView):
    """
    Chilean honorary tax calculator for 2025.

    POST /api/v1/billing/tax-calculator/
    Permission: IsDoctor

    Accepts: {"gross_amount": 500000, "document_type": "boleta"}
    document_type: "boleta" | "factura_exenta"

    Always returns BOTH calculations (boleta + factura) for side-by-side comparison.

    Chilean 2025 tax parameters:
      - Boleta de Honorarios: 13.75% retention at source (PPM provisional)
      - Factura Exenta: no withholding; doctor pays annual impuesto global complementario
      - PPM provisional estimate: 10% of gross (applies to both)
      - UTM 2025 ≈ 67,000 CLP

    Progressive tax brackets (monthly, UTM-based):
      0–13.5 UTM: 0% | 13.5–30 UTM: 4% | 30–50 UTM: 8% | 50–70 UTM: 13.5%
      70–90 UTM: 23% | 90–120 UTM: 30.4% | 120–310 UTM: 35% | >310 UTM: 40%
    """

    permission_classes = [IsDoctor]

    # 2025 constants
    RETENTION_RATE = 0.1375  # boleta de honorarios withholding
    PPM_RATE = 0.10           # provisional monthly payment estimate
    UTM_2025 = 67_000         # CLP per UTM, approximate

    # (upper_utm_limit, marginal_rate) — last entry has no upper limit (None)
    TAX_BRACKETS = [
        (13.5,  0.00),
        (30.0,  0.04),
        (50.0,  0.08),
        (70.0,  0.135),
        (90.0,  0.23),
        (120.0, 0.304),
        (310.0, 0.35),
        (None,  0.40),
    ]

    def _effective_annual_tax(self, gross_monthly: float) -> float:
        """
        Estimate annual income tax for a factura_exenta issuer.

        Applies the Chilean progressive brackets to the monthly gross,
        multiplies by 12 for the annual figure.
        Monthly gross is converted to UTM for bracket comparison.
        """
        monthly_utm = gross_monthly / self.UTM_2025
        prev_limit = 0.0
        monthly_tax = 0.0

        for upper, rate in self.TAX_BRACKETS:
            if rate == 0.0:
                prev_limit = upper if upper is not None else monthly_utm
                continue
            upper_utm = upper if upper is not None else float("inf")
            taxable_utm = max(0.0, min(monthly_utm, upper_utm) - prev_limit)
            monthly_tax += taxable_utm * self.UTM_2025 * rate
            if monthly_utm <= upper_utm:
                break
            prev_limit = upper_utm

        return monthly_tax * 12

    def post(self, request: Request) -> Response:
        """Calculate and compare boleta vs factura_exenta tax scenarios."""
        gross_amount = request.data.get("gross_amount")
        document_type = request.data.get("document_type", "boleta")

        if gross_amount is None:
            return Response(
                {"detail": "gross_amount is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            gross = float(gross_amount)
        except (TypeError, ValueError):
            return Response(
                {"detail": "gross_amount must be a number."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if gross <= 0:
            return Response(
                {"detail": "gross_amount must be greater than zero."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if document_type not in ("boleta", "factura_exenta"):
            return Response(
                {"detail": "document_type must be 'boleta' or 'factura_exenta'."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # ── Boleta de Honorarios ───────────────────────────────────────────────
        boleta_retention = round(gross * self.RETENTION_RATE)
        boleta_net = round(gross - boleta_retention)
        boleta_ppm = round(gross * self.PPM_RATE)

        # ── Factura Exenta ─────────────────────────────────────────────────────
        factura_annual_tax = round(self._effective_annual_tax(gross))
        factura_ppm = round(gross * self.PPM_RATE)

        # ── Annual projections (12 months) ─────────────────────────────────────
        boleta_annual_net = boleta_net * 12
        # For factura: net after annual tax obligation (divided evenly per month * 12)
        factura_annual_net = round(gross * 12 - factura_annual_tax)
        difference = abs(factura_annual_net - boleta_annual_net)
        recommendation = "factura" if factura_annual_net >= boleta_annual_net else "boleta"

        return Response(
            {
                "gross_amount": gross,
                "document_type": document_type,
                "boleta": {
                    "retention_rate": self.RETENTION_RATE,
                    "retention": boleta_retention,
                    "net_amount": boleta_net,
                    "monthly_ppm": boleta_ppm,
                },
                "factura": {
                    "retention": 0,
                    "net_amount": round(gross),
                    "monthly_ppm": factura_ppm,
                    "annual_tax_estimate": factura_annual_tax,
                },
                "comparison": {
                    "boleta_annual_net": boleta_annual_net,
                    "factura_annual_net": factura_annual_net,
                    "difference": difference,
                    "recommendation": recommendation,
                },
            },
            status=status.HTTP_200_OK,
        )


class PaymentProviderViewSet(viewsets.ModelViewSet):
    """
    CRUD for PaymentProvider configuration (doctor admin only).

    Allows the doctor to configure MercadoPago credentials per practice.
    """

    serializer_class = PaymentProviderSerializer
    permission_classes = [IsDoctor]
    pagination_class = StandardPagination
    http_method_names = ["get", "post", "put", "patch", "delete", "head", "options"]

    def get_queryset(self):
        from apps.practice.models import Practice

        # Doctor sees only their own practice's providers
        try:
            practice = Practice.objects.get(owner=self.request.user)
            return PaymentProvider.objects.filter(practice=practice)
        except Practice.DoesNotExist:
            return PaymentProvider.objects.none()


class MonthlyExpenseViewSet(viewsets.ModelViewSet):
    """
    CRUD for recurring monthly expenses (doctor only).

    The practice is resolved automatically from the authenticated doctor —
    it does not need to be sent in the request body.
    Supports filtering by is_active and category via query params.
    """

    serializer_class = MonthlyExpenseSerializer
    permission_classes = [IsDoctor]
    pagination_class = StandardPagination
    http_method_names = ["get", "post", "put", "patch", "delete", "head", "options"]

    def _get_practice(self):
        from apps.practice.models import Practice

        return Practice.objects.get(owner=self.request.user)

    def get_queryset(self):
        try:
            practice = self._get_practice()
        except Exception:
            return MonthlyExpense.objects.none()

        qs = MonthlyExpense.objects.filter(practice=practice)

        is_active_param = self.request.query_params.get("is_active")
        if is_active_param is not None:
            qs = qs.filter(is_active=is_active_param.lower() == "true")

        category_param = self.request.query_params.get("category")
        if category_param:
            qs = qs.filter(category=category_param)

        return qs

    def perform_create(self, serializer):
        practice = self._get_practice()
        serializer.save(practice=practice)


class TaxSummaryView(APIView):
    """
    GET /api/v1/billing/tax-summary/

    Returns a tax/financial summary for the given period. Designed to give
    the accountant a quick snapshot of gross income, retention estimate, and
    breakdowns by payment method, service, and insurance.

    Query params:
      ?period=monthly&year=2025&month=6   (default: current month)
      ?period=quarterly&year=2025&quarter=2
      ?period=annual&year=2025

    Permission: IsDoctor
    """

    permission_classes = [IsDoctor]

    RETENTION_RATE = Decimal("0.1375")  # boleta de honorarios 13.75%

    # Spanish month names for labels
    _MONTH_NAMES = [
        "", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
        "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
    ]

    def _parse_period(self, params: dict, today: date) -> tuple[date, date, dict]:
        """
        Parse period query params and return (start_date, end_date, period_meta).

        Raises ValueError with a human-readable message on invalid input.
        """
        from calendar import monthrange

        period_type = params.get("period", "monthly")

        try:
            year = int(params.get("year", today.year))
        except (TypeError, ValueError):
            raise ValueError("year must be an integer.")

        if period_type == "monthly":
            try:
                month = int(params.get("month", today.month))
            except (TypeError, ValueError):
                raise ValueError("month must be an integer.")
            if not 1 <= month <= 12:
                raise ValueError("month must be between 1 and 12.")
            _, last_day = monthrange(year, month)
            start = date(year, month, 1)
            end = date(year, month, last_day)
            label = f"{self._MONTH_NAMES[month]} {year}"
            meta = {"type": "monthly", "year": year, "month": month, "label": label}

        elif period_type == "quarterly":
            try:
                quarter = int(params.get("quarter", 1))
            except (TypeError, ValueError):
                raise ValueError("quarter must be an integer.")
            if not 1 <= quarter <= 4:
                raise ValueError("quarter must be between 1 and 4.")
            first_month = (quarter - 1) * 3 + 1
            last_month = first_month + 2
            _, last_day = monthrange(year, last_month)
            start = date(year, first_month, 1)
            end = date(year, last_month, last_day)
            label = f"T{quarter} {year}"
            meta = {"type": "quarterly", "year": year, "quarter": quarter, "label": label}

        elif period_type == "annual":
            start = date(year, 1, 1)
            end = date(year, 12, 31)
            label = str(year)
            meta = {"type": "annual", "year": year, "label": label}

        else:
            raise ValueError("period must be 'monthly', 'quarterly', or 'annual'.")

        return start, end, meta

    def _build_monthly_trend(self, practice, today: date) -> list[dict]:
        """Return last 12 months of completed payment totals (inclusive of current month)."""
        from django.db.models import Count

        months: list[tuple[int, int]] = []
        year, month = today.year, today.month
        for _ in range(12):
            months.insert(0, (year, month))
            month -= 1
            if month == 0:
                month = 12
                year -= 1

        trend_qs = (
            Payment.objects.filter(
                practice=practice,
                status=Payment.COMPLETED,
                paid_at__isnull=False,
            )
            .annotate(month=TruncMonth("paid_at"))
            .values("month")
            .annotate(income=Sum("amount"), count=Count("id"))
            .order_by("month")
        )

        trend_by_month: dict[str, dict] = {}
        for row in trend_qs:
            key = row["month"].strftime("%Y-%m")
            trend_by_month[key] = {"income": int(row["income"]), "count": row["count"]}

        result = []
        for y, m in months:
            key = f"{y:04d}-{m:02d}"
            entry = trend_by_month.get(key, {"income": 0, "count": 0})
            result.append({"month": key, "income": entry["income"], "count": entry["count"]})
        return result

    def get(self, request: Request) -> Response:
        """Return tax summary for the requested period."""
        from calendar import monthrange
        from django.db.models import Count, Q

        from apps.practice.models import Practice

        try:
            practice = Practice.objects.get(owner=request.user)
        except Practice.DoesNotExist:
            return Response(
                {"detail": "No practice found for this user."},
                status=status.HTTP_404_NOT_FOUND,
            )

        today = date.today()

        try:
            start, end, period_meta = self._parse_period(request.query_params, today)
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        # ── Base queryset: completed payments in period ────────────────────────
        base_qs = Payment.objects.filter(
            practice=practice,
            status=Payment.COMPLETED,
            paid_at__date__gte=start,
            paid_at__date__lte=end,
        )

        # ── Summary ───────────────────────────────────────────────────────────
        agg = base_qs.aggregate(gross=Sum("amount"), count=Count("id"))
        gross_income = int(agg["gross"] or 0)
        completed_count = agg["count"] or 0
        avg_per_appointment = round(gross_income / completed_count) if completed_count else 0
        retention_estimate = int(Decimal(gross_income) * self.RETENTION_RATE)
        net_after_retention = gross_income - retention_estimate

        # ── By payment method ─────────────────────────────────────────────────
        method_rows = (
            base_qs
            .values("payment_method")
            .annotate(count=Count("id"), total=Sum("amount"))
            .order_by("-total")
        )
        by_payment_method = [
            {
                "method": row["payment_method"],
                "count": row["count"],
                "total": int(row["total"]),
                "percentage": round(int(row["total"]) / gross_income * 100, 1) if gross_income else 0.0,
            }
            for row in method_rows
        ]

        # ── By service ────────────────────────────────────────────────────────
        service_rows = (
            base_qs
            .filter(appointment__service__isnull=False)
            .values("appointment__service__name")
            .annotate(count=Count("id"), total=Sum("amount"))
            .order_by("-total")
        )
        by_service = [
            {
                "service_name": row["appointment__service__name"],
                "count": row["count"],
                "total": int(row["total"]),
                "percentage": round(int(row["total"]) / gross_income * 100, 1) if gross_income else 0.0,
            }
            for row in service_rows
        ]

        # ── By insurance ──────────────────────────────────────────────────────
        insurance_rows = (
            base_qs
            .values("patient__insurance")
            .annotate(count=Count("id"), total=Sum("amount"))
            .order_by("-total")
        )
        by_insurance = [
            {
                "insurance": row["patient__insurance"] or "",
                "count": row["count"],
                "total": int(row["total"]),
                "percentage": round(int(row["total"]) / gross_income * 100, 1) if gross_income else 0.0,
            }
            for row in insurance_rows
        ]

        # ── Monthly trend (last 12 months, always) ────────────────────────────
        monthly_trend = self._build_monthly_trend(practice, today)

        return Response(
            {
                "period": period_meta,
                "summary": {
                    "gross_income": gross_income,
                    "completed_payments": completed_count,
                    "avg_per_appointment": avg_per_appointment,
                    "retention_estimate": retention_estimate,
                    "net_after_retention": net_after_retention,
                },
                "by_payment_method": by_payment_method,
                "by_service": by_service,
                "by_insurance": by_insurance,
                "monthly_trend": monthly_trend,
            },
            status=status.HTTP_200_OK,
        )


class TaxSummaryExportView(APIView):
    """
    GET /api/v1/billing/tax-summary/export/

    Exports the tax summary for the given period as an XLSX file with 4 sheets.
    Accepts the same period query params as TaxSummaryView.

    Sheets (Spanish headers):
      - Resumen: period, gross, retention, net, count, avg
      - Por Método de Pago: method, count, total, %
      - Por Servicio: service, count, total, %
      - Por Previsión: insurance, count, total, %

    Permission: IsDoctor
    """

    permission_classes = [IsDoctor]

    def get(self, request: Request) -> HttpResponse:
        """Generate and return the XLSX summary file."""
        from apps.practice.models import Practice

        try:
            practice = Practice.objects.get(owner=request.user)
        except Practice.DoesNotExist:
            return Response(
                {"detail": "No practice found for this user."},
                status=status.HTTP_404_NOT_FOUND,
            )

        # Reuse TaxSummaryView parsing logic by instantiating it
        summary_view = TaxSummaryView()
        today = date.today()
        try:
            start, end, period_meta = summary_view._parse_period(request.query_params, today)
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        base_qs = Payment.objects.filter(
            practice=practice,
            status=Payment.COMPLETED,
            paid_at__date__gte=start,
            paid_at__date__lte=end,
        )

        from django.db.models import Count

        agg = base_qs.aggregate(gross=Sum("amount"), count=Count("id"))
        gross_income = int(agg["gross"] or 0)
        completed_count = agg["count"] or 0
        avg_per_appointment = round(gross_income / completed_count) if completed_count else 0
        retention_estimate = int(Decimal(gross_income) * TaxSummaryView.RETENTION_RATE)
        net_after_retention = gross_income - retention_estimate

        method_rows = list(
            base_qs
            .values("payment_method")
            .annotate(count=Count("id"), total=Sum("amount"))
            .order_by("-total")
        )
        service_rows = list(
            base_qs
            .filter(appointment__service__isnull=False)
            .values("appointment__service__name")
            .annotate(count=Count("id"), total=Sum("amount"))
            .order_by("-total")
        )
        insurance_rows = list(
            base_qs
            .values("patient__insurance")
            .annotate(count=Count("id"), total=Sum("amount"))
            .order_by("-total")
        )

        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            return Response(
                {"detail": "openpyxl is not installed. Cannot export XLSX."},
                status=status.HTTP_501_NOT_IMPLEMENTED,
            )

        # ── Styles ────────────────────────────────────────────────────────────
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        label_font = Font(bold=True)

        def apply_header(ws, headers: list[str]) -> None:
            for col_idx, h in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=h)
                cell.font = header_font
                cell.fill = header_fill
            for col_idx in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = 22

        def pct(total: int, gross: int) -> float:
            return round(total / gross * 100, 1) if gross else 0.0

        wb = openpyxl.Workbook()

        # ── Sheet 1: Resumen ──────────────────────────────────────────────────
        ws_resumen = wb.active
        ws_resumen.title = "Resumen"

        summary_rows = [
            ("Período", period_meta["label"]),
            ("Tipo de período", period_meta["type"].capitalize()),
            ("Ingresos brutos (CLP)", gross_income),
            ("Pagos completados", completed_count),
            ("Promedio por consulta (CLP)", avg_per_appointment),
            ("Retención estimada 13.75% (CLP)", retention_estimate),
            ("Neto tras retención (CLP)", net_after_retention),
        ]
        ws_resumen.column_dimensions["A"].width = 34
        ws_resumen.column_dimensions["B"].width = 22
        for row_idx, (label, value) in enumerate(summary_rows, start=1):
            cell_label = ws_resumen.cell(row=row_idx, column=1, value=label)
            cell_label.font = label_font
            ws_resumen.cell(row=row_idx, column=2, value=value)

        # ── Sheet 2: Por Método de Pago ───────────────────────────────────────
        ws_method = wb.create_sheet("Por Método de Pago")
        apply_header(ws_method, ["Método de Pago", "Cantidad", "Total (CLP)", "Porcentaje (%)"])
        for row_idx, row in enumerate(method_rows, start=2):
            ws_method.cell(row=row_idx, column=1, value=row["payment_method"])
            ws_method.cell(row=row_idx, column=2, value=row["count"])
            ws_method.cell(row=row_idx, column=3, value=int(row["total"]))
            ws_method.cell(row=row_idx, column=4, value=pct(int(row["total"]), gross_income))

        # ── Sheet 3: Por Servicio ─────────────────────────────────────────────
        ws_service = wb.create_sheet("Por Servicio")
        apply_header(ws_service, ["Servicio", "Cantidad", "Total (CLP)", "Porcentaje (%)"])
        for row_idx, row in enumerate(service_rows, start=2):
            ws_service.cell(row=row_idx, column=1, value=row["appointment__service__name"])
            ws_service.cell(row=row_idx, column=2, value=row["count"])
            ws_service.cell(row=row_idx, column=3, value=int(row["total"]))
            ws_service.cell(row=row_idx, column=4, value=pct(int(row["total"]), gross_income))

        # ── Sheet 4: Por Previsión ────────────────────────────────────────────
        ws_insurance = wb.create_sheet("Por Previsión")
        apply_header(ws_insurance, ["Previsión", "Cantidad", "Total (CLP)", "Porcentaje (%)"])
        for row_idx, row in enumerate(insurance_rows, start=2):
            ws_insurance.cell(row=row_idx, column=1, value=row["patient__insurance"] or "Sin especificar")
            ws_insurance.cell(row=row_idx, column=2, value=row["count"])
            ws_insurance.cell(row=row_idx, column=3, value=int(row["total"]))
            ws_insurance.cell(row=row_idx, column=4, value=pct(int(row["total"]), gross_income))

        # ── Serialize ─────────────────────────────────────────────────────────
        import io as _io
        buffer = _io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        period_label = period_meta["label"].replace(" ", "_")
        filename = f"resumen_tributario_{period_label}.xlsx"

        response = HttpResponse(
            content=buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


class CashFlowView(APIView):
    """
    GET /api/v1/billing/cash-flow/

    Returns income (completed payments) vs expenses (active monthly expenses)
    for the last 6 months, plus a breakdown for the current month.
    """

    permission_classes = [IsDoctor]

    def get(self, request: Request) -> Response:
        from apps.practice.models import Practice

        try:
            practice = Practice.objects.get(owner=request.user)
        except Practice.DoesNotExist:
            return Response(
                {"detail": "No practice found for this user."},
                status=status.HTTP_404_NOT_FOUND,
            )

        today = date.today()

        # Build the last 6 months as (year, month) tuples, chronological
        months: list[tuple[int, int]] = []
        year, month = today.year, today.month
        for _ in range(6):
            months.insert(0, (year, month))
            month -= 1
            if month == 0:
                month = 12
                year -= 1

        # ── Income: aggregate completed payments per month ────────────────────
        from django.db.models import Q

        income_qs = (
            Payment.objects.filter(
                practice=practice,
                status=Payment.COMPLETED,
                paid_at__isnull=False,
            )
            .annotate(month=TruncMonth("paid_at"))
            .values("month")
            .annotate(total=Sum("amount"))
            .order_by("month")
        )

        income_by_month: dict[str, int] = {}
        for row in income_qs:
            key = row["month"].strftime("%Y-%m")
            income_by_month[key] = int(row["total"])

        # ── Expenses: active MonthlyExpense is a fixed monthly cost ───────────
        active_expenses = list(
            MonthlyExpense.objects.filter(practice=practice, is_active=True)
        )
        total_expenses = sum(e.amount for e in active_expenses)

        # ── Build per-month series ─────────────────────────────────────────────
        month_series = []
        for y, m in months:
            key = f"{y:04d}-{m:02d}"
            income = income_by_month.get(key, 0)
            month_series.append(
                {
                    "month": key,
                    "income": income,
                    "expenses": total_expenses,
                    "net": income - total_expenses,
                }
            )

        # ── Current month detail ───────────────────────────────────────────────
        current_key = f"{today.year:04d}-{today.month:02d}"
        current_income = income_by_month.get(current_key, 0)
        expenses_breakdown = [
            {"name": e.name, "category": e.category, "amount": e.amount}
            for e in sorted(active_expenses, key=lambda x: -x.amount)
        ]

        return Response(
            {
                "months": month_series,
                "current_month": {
                    "income": current_income,
                    "total_expenses": total_expenses,
                    "net": current_income - total_expenses,
                    "expenses_breakdown": expenses_breakdown,
                },
            },
            status=status.HTTP_200_OK,
        )
